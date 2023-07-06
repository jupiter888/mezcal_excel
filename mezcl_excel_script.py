#This code will create an Excel workbook named "mezcal_notebook" with the specified worksheets and column headers. 
#Once created, this script will save the workbook as an Excel file called "mezcal_notebook.xlsx".


#explained script structure:
#python code to create creating an excel notebook named mezcal_notebook, and create a set of 5 worksheets,
#generate a unique key id for each of the rows primary data(the restaurant_name is the primary key) and the secondary key is assigned to each of the restaurants for reference to each in the other tables on each of the 5 worksheets. 

#worksheet 1 named contacts_info with columns: restaurant_name, manager_name, contact_phone, contact_email, optimal_time, restaurant_address, and restaurant_id. 
#worksheet 2 named inside_info: current_mezcals, desired_mezcals,sale_price(null values default),  tasting_required, tasted_bool, preferred_bottles, restaurant_id. 
#worksheet 3 tastings_info: restaurant_id, interested_bool, contacted_again_bool, tasting_scheduled_bool, tasting_date. 
#worksheet 4 named contact_notes: restaurant_id, notes. 
#worksheet 5 named bottles_info: restaurant_id, bottles_sold, quantity_sold, bottle_price, cases_sold_bool, case_deal_bool, client_potential_scale. save the excel. 


from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Create a new workbook
workbook = Workbook()

# Rename the default sheet to "contacts_info"
worksheet1 = workbook.active
worksheet1.title = "contacts_info"

# Define the column headers for worksheet 1
headers1 = ["restaurant_name", "manager_name", "contact_phone", "contact_email", "optimal_time", "restaurant_address", "restaurant_id"]

# Write the column headers to worksheet 1
for col_num, header in enumerate(headers1, 1):
    col_letter = get_column_letter(col_num)
    worksheet1[f"{col_letter}1"] = header

# Rename the sheet to "inside_info"
worksheet2 = workbook.create_sheet(title="inside_info")

# Define the column headers for worksheet 2
headers2 = ["current_mezcals", "desired_mezcals", "sale_price", "tasting_required", "tasted_bool", "preferred_bottles", "restaurant_id"]

# Write the column headers to worksheet 2
for col_num, header in enumerate(headers2, 1):
    col_letter = get_column_letter(col_num)
    worksheet2[f"{col_letter}1"] = header

# Rename the sheet to "tastings_info"
worksheet3 = workbook.create_sheet(title="tastings_info")

# Define the column headers for worksheet 3
headers3 = ["restaurant_id", "interested_bool", "contacted_again_bool", "tasting_scheduled_bool", "tasting_date"]

# Write the column headers to worksheet 3
for col_num, header in enumerate(headers3, 1):
    col_letter = get_column_letter(col_num)
    worksheet3[f"{col_letter}1"] = header

# Rename the sheet to "contact_notes"
worksheet4 = workbook.create_sheet(title="contact_notes")

# Define the column headers for worksheet 4
headers4 = ["restaurant_id", "notes"]

# Write the column headers to worksheet 4
for col_num, header in enumerate(headers4, 1):
    col_letter = get_column_letter(col_num)
    worksheet4[f"{col_letter}1"] = header

# Rename the sheet to "bottles_info"
worksheet5 = workbook.create_sheet(title="bottles_info")

# Define the column headers for worksheet 5
headers5 = ["restaurant_id", "bottles_sold", "quantity_sold", "bottle_price", "cases_sold_bool", "case_deal_bool", "client_potential_scale"]

# Write the column headers to worksheet 5
for col_num, header in enumerate(headers5, 1):
    col_letter = get_column_letter(col_num)
    worksheet5[f"{col_letter}1"] = header

# Save the workbook
workbook.save("mezcal_notebook.xlsx")
